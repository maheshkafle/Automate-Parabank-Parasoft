import openpyxl
import xlsxwriter

def filecreate(filename):
    workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet()
    workbook.close()

def getrowcount(filepath,sheetname):
    Workbook = openpyxl.load_workbook(filepath)
    Sheet = Workbook.get_sheet_by_name(sheetname)
    Rw = Sheet.max_row
    return Rw

def getcolumnount(filepath,sheetname):
    Workbook = openpyxl.load_workbook(filepath)
    Sheet = Workbook.get_sheet_by_name(sheetname)
    Cln = Sheet.max_column
    return Cln

def readfromxl(filepath,sheetname,rownum,colnum):
    Workbook = openpyxl.load_workbook(filepath)
    Sheet = Workbook.get_sheet_by_name(sheetname)
    return Sheet.cell(rownum,colnum).value

def writetoxl(filepath,sheetname,rownum,colnum,data):
    Workbook = openpyxl.load_workbook(filepath)
    Sheet = Workbook.get_sheet_by_name(sheetname)
    Sheet.cell(rownum,colnum).value = data
    Workbook.save(filepath)